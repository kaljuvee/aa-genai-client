import os
import json
import logging
from azure.core.credentials import AzureKeyCredential
from azure.search.documents import SearchClient
from azure.search.documents.indexes import SearchIndexClient
from azure.search.documents.indexes.models import (
    SearchIndex,
    SimpleField,
    SearchableField,
    SearchFieldDataType,
    CustomAnalyzer,
)
from dotenv import load_dotenv
import base64
import PyPDF2
import openpyxl
import pandas as pd
import time

# Load environment variables
load_dotenv()

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Global variables
DOCUMENT_DIR = "docs"
CONFIG_FILE = "config/index_metadata_multi.json"

# Azure AI Search configuration
search_endpoint = os.getenv("AZURE_SEARCH_ENDPOINT")
search_key = os.getenv("AZURE_SEARCH_KEY")
logging.info(f"Azure Search Endpoint: {search_endpoint}")
logging.info(f"Azure Search Key: {'*' * len(search_key) if search_key else 'Not found'}")

try:
    search_credential = AzureKeyCredential(search_key)
    logging.info("Azure credentials created successfully")
except ValueError as e:
    logging.error(f"Error creating Azure credentials: {str(e)}")
    raise

try:
    # Initialize clients
    search_index_client = SearchIndexClient(endpoint=search_endpoint, credential=search_credential, api_version="2023-10-01-Preview")
    logging.info("Azure index client initialized successfully")
except Exception as e:
    logging.error(f"Error initializing Azure index client: {str(e)}")
    raise

def create_search_index(index_name):
    logging.info(f"Creating search index: {index_name}")
    try:
        # Define a custom analyzer
        custom_analyzer = CustomAnalyzer(
            name="custom_analyzer",
            tokenizer_name="microsoft_language_tokenizer",
            token_filters=["lowercase"]
        )

        fields = [
            SimpleField(name="id", type=SearchFieldDataType.String, key=True),
            SearchableField(
                name="content",
                type=SearchFieldDataType.String,
                analyzer_name="custom_analyzer"
            ),
        ]
        
        # Create the index with the custom analyzer
        index = SearchIndex(
            name=index_name,
            fields=fields,
            analyzers=[custom_analyzer]
        )
        result = search_index_client.create_or_update_index(index)
        logging.info(f"Search index '{index_name}' created successfully. Result: {result}")
    except Exception as e:
        logging.error(f"Error creating search index: {str(e)}")
        raise

def encode_filename(filename):
    # Remove the .pdf extension, encode to bytes, then to base64, and decode to string
    return base64.urlsafe_b64encode(filename[:-4].encode()).decode()

def read_pdf(file_path):
    logging.info(f"Reading PDF file: {file_path}")
    with open(file_path, 'rb') as file:
        pdf_reader = PyPDF2.PdfReader(file)
        content = ""
        for page in pdf_reader.pages:
            content += page.extract_text()
    return content.strip()

def read_xlsx(file_path):
    logging.info(f"Reading XLSX file: {file_path}")
    workbook = openpyxl.load_workbook(file_path)
    content = ""
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        for row in worksheet.iter_rows(values_only=True):
            content += " ".join(str(cell) for cell in row if cell is not None) + "\n"
    return content.strip()

def read_csv(file_path):
    logging.info(f"Reading CSV file: {file_path}")
    df = pd.read_csv(file_path)
    content = df.to_string()
    return content.strip()

# Add this new function to detect document type
def get_document_type(filename):
    """Detect document type from file extension."""
    extension = filename.lower().split('.')[-1]
    if extension in ['pdf', 'xlsx', 'csv']:
        return extension
    else:
        raise ValueError(f"Unsupported file extension: {extension}")

# Modify read_and_index_document to handle document type detection
def read_and_index_document(index_name, document_file):
    logging.info(f"Reading and indexing document for index {index_name}")
    try:
        file_path = os.path.join(DOCUMENT_DIR, document_file.strip())
        logging.info(f"Reading file: {file_path}")
        
        document_type = get_document_type(document_file)
        
        if document_type == 'pdf':
            content = read_pdf(file_path)
        elif document_type == 'xlsx':
            content = read_xlsx(file_path)
        elif document_type == 'csv':
            content = read_csv(file_path)
        else:
            raise ValueError(f"Unsupported document type: {document_type}")
        
        encoded_filename = encode_filename(document_file)
        document = {
            "id": encoded_filename,
            "content": content
        }
        
        logging.info(f"Successfully read file: {document_file} (encoded as: {encoded_filename})")
    
        # Initialize SearchClient for this specific index
        search_client = SearchClient(endpoint=search_endpoint, 
                                   index_name=index_name, 
                                   credential=search_credential, 
                                   api_version="2023-10-01-Preview")
        
        logging.info(f"Attempting to index document for {index_name}")
        result = search_client.upload_documents([document])
        succeeded = sum(1 for r in result if r.succeeded)
        failed = sum(1 for r in result if not r.succeeded)
        logging.info(f"Indexing complete for {index_name}. Succeeded: {succeeded}, Failed: {failed}")
        if failed > 0:
            logging.warning(f"Document failed to index for {index_name}. Check individual results for details.")
        
        return search_client
    except Exception as e:
        logging.error(f"Error in read_and_index_document for {index_name}: {str(e)}")
        raise

def delete_index_if_exists(index_name):
    try:
        search_index_client.delete_index(index_name)
        logging.info(f"Existing index '{index_name}' deleted successfully.")
    except Exception as e:
        if "ResourceNotFound" not in str(e):
            logging.error(f"Error deleting index {index_name}: {str(e)}")
            raise
        else:
            logging.info(f"Index '{index_name}' does not exist. Proceeding with creation.")

def main():
    logging.info("Starting the index creation and document upload process")

    try:
        # Read the configuration file
        with open(CONFIG_FILE, 'r') as config_file:
            index_config = json.load(config_file)
        
        for _, index_data in index_config.items():
            index_name = index_data['index_name']
            document_list = [doc.strip() for doc in index_data['document_list'].split(',')]
            logging.info(f"Processing index: {index_name} with documents: {document_list}")

            # Delete existing index if it exists
            delete_index_if_exists(index_name)

            # Create search index
            create_search_index(index_name)

            # Read and index each document
            search_client = None
            for document_file in document_list:
                search_client = read_and_index_document(index_name, document_file)

            # Check final document count
            if search_client:
                for i in range(5):
                    total_docs = search_client.get_document_count()
                    logging.info(f"Attempt {i+1}: Total documents in index {index_name}: {total_docs}")
                    if total_docs == len(document_list):
                        break
                    time.sleep(5)

                if total_docs != len(document_list):
                    logging.warning(f"Expected {len(document_list)} documents in index {index_name}, but found {total_docs}")

        logging.info("Index creation and document upload complete for all indices.")
    except Exception as e:
        logging.error(f"An error occurred during the main process: {str(e)}")
        print(f"An error occurred. Please check the logs for details.")

if __name__ == "__main__":
    main()
